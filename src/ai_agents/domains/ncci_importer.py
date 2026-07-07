"""CMS NCCI PTP edit file importer.

The importer accepts CMS ZIP/TXT/CSV files downloaded from the Medicare NCCI
Procedure-to-Procedure edits page and loads active PTP edits into the configured
SQLite reference database. It is intentionally streaming and file-format tolerant
because CMS file headers vary slightly across releases.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import sqlite3
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


class NCCIImportError(RuntimeError):
    """Raised when NCCI files cannot be imported."""


@dataclass(frozen=True)
class NCCIImportSummary:
    """Summary of imported NCCI rows."""

    files_seen: int
    rows_seen: int
    rows_imported: int
    rows_skipped: int

    def to_dict(self) -> dict[str, int]:
        return {
            "files_seen": self.files_seen,
            "rows_seen": self.rows_seen,
            "rows_imported": self.rows_imported,
            "rows_skipped": self.rows_skipped,
        }


@dataclass(frozen=True)
class NCCIPTPRow:
    """Normalized NCCI PTP edit row."""

    code_a: str
    code_b: str
    modifier_indicator: str
    effective_date: str | None
    deletion_date: str | None
    rationale: str | None
    source_file: str


def import_ncci_ptp_files(
    *,
    db_path: str | Path,
    files: Iterable[str | Path],
    edit_type: str,
    import_version: str,
) -> NCCIImportSummary:
    """Import CMS NCCI PTP edit files into SQLite."""

    rows = _active_rows_from_files(files)
    with sqlite3.connect(Path(db_path)) as conn:
        for row in rows.active:
            conn.execute(
                """
                INSERT OR REPLACE INTO ncci_ptp_edits (
                    code_a,
                    code_b,
                    modifier_indicator,
                    edit_type,
                    effective_date,
                    deletion_date,
                    rationale,
                    source_file,
                    import_version
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row.code_a,
                    row.code_b,
                    row.modifier_indicator,
                    edit_type,
                    row.effective_date,
                    row.deletion_date,
                    row.rationale,
                    row.source_file,
                    import_version,
                ),
            )
        conn.commit()

    return rows.summary


def import_ncci_ptp_files_to_postgres(
    *,
    postgres_url: str,
    files: Iterable[str | Path],
    edit_type: str,
    import_version: str,
) -> NCCIImportSummary:
    """Import CMS NCCI PTP edit files into Supabase/Postgres."""

    rows = _active_rows_from_files(files)
    try:
        import psycopg
    except ImportError as exc:
        raise NCCIImportError(
            'Postgres import requires: python -m pip install -e ".[postgres]"'
        ) from exc

    with psycopg.connect(postgres_url) as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """
                insert into public.ncci_ptp_edits (
                    code_a,
                    code_b,
                    modifier_indicator,
                    edit_type,
                    effective_date,
                    deletion_date,
                    rationale,
                    source_file,
                    import_version
                ) values (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                on conflict (code_a, code_b, edit_type) do update set
                    modifier_indicator = excluded.modifier_indicator,
                    effective_date = excluded.effective_date,
                    deletion_date = excluded.deletion_date,
                    rationale = excluded.rationale,
                    source_file = excluded.source_file,
                    import_version = excluded.import_version
                """,
                [
                    (
                        row.code_a,
                        row.code_b,
                        row.modifier_indicator,
                        edit_type,
                        row.effective_date,
                        row.deletion_date,
                        row.rationale,
                        row.source_file,
                        import_version,
                    )
                    for row in rows.active
                ],
            )
        conn.commit()

    return rows.summary


def load_normalized_ncci_csv_to_sqlite(
    *,
    db_path: str | Path,
    csv_files: Iterable[str | Path],
    batch_size: int = 10_000,
) -> NCCIImportSummary:
    """Load already-normalized NCCI CSV files into SQLite in batches."""

    files_seen = 0
    rows_seen = 0
    rows_imported = 0
    rows_skipped = 0
    db = Path(db_path)
    db.parent.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(db) as conn:
        conn.execute("pragma journal_mode = wal")
        conn.execute("pragma synchronous = normal")
        conn.execute("pragma temp_store = memory")
        for csv_file in csv_files:
            files_seen += 1
            batch: list[tuple[str, str, str, str, str | None, str | None, str | None, str | None, str | None]] = []
            with Path(csv_file).open("r", encoding="utf-8", newline="") as handle:
                reader = csv.DictReader(handle)
                for row in reader:
                    rows_seen += 1
                    modifier_indicator = row.get("modifier_indicator", "").strip()
                    if modifier_indicator not in {"0", "1"}:
                        rows_skipped += 1
                        continue
                    batch.append(
                        (
                            row.get("code_a", "").strip(),
                            row.get("code_b", "").strip(),
                            modifier_indicator,
                            row.get("edit_type", "").strip(),
                            _blank_to_none(row.get("effective_date")),
                            _blank_to_none(row.get("deletion_date")),
                            _blank_to_none(row.get("rationale")),
                            _blank_to_none(row.get("source_file")),
                            _blank_to_none(row.get("import_version")),
                        )
                    )
                    if len(batch) >= batch_size:
                        rows_imported += _insert_sqlite_batch(conn, batch)
                        batch.clear()
                if batch:
                    rows_imported += _insert_sqlite_batch(conn, batch)
        conn.commit()

    return NCCIImportSummary(
        files_seen=files_seen,
        rows_seen=rows_seen,
        rows_imported=rows_imported,
        rows_skipped=rows_skipped,
    )


def _insert_sqlite_batch(
    conn: sqlite3.Connection,
    batch: list[tuple[str, str, str, str, str | None, str | None, str | None, str | None, str | None]],
) -> int:
    conn.executemany(
        """
        INSERT OR REPLACE INTO ncci_ptp_edits (
            code_a,
            code_b,
            modifier_indicator,
            edit_type,
            effective_date,
            deletion_date,
            rationale,
            source_file,
            import_version
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        batch,
    )
    return len(batch)


def extract_ncci_ptp_files_to_csv(
    *,
    files: Iterable[str | Path],
    output_csv: str | Path,
    edit_type: str,
    import_version: str,
) -> NCCIImportSummary:
    """Extract CMS NCCI PTP source files into a normalized CSV."""

    rows = _active_rows_from_files(files)
    output_path = Path(output_csv)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "code_a",
                "code_b",
                "modifier_indicator",
                "edit_type",
                "effective_date",
                "deletion_date",
                "rationale",
                "source_file",
                "import_version",
            ]
        )
        for row in rows.active:
            writer.writerow(
                [
                    row.code_a,
                    row.code_b,
                    row.modifier_indicator,
                    edit_type,
                    row.effective_date or "",
                    row.deletion_date or "",
                    row.rationale or "",
                    row.source_file,
                    import_version,
                ]
            )
    return rows.summary


@dataclass(frozen=True)
class _RowsFromFiles:
    active: tuple[NCCIPTPRow, ...]
    summary: NCCIImportSummary


def _active_rows_from_files(files: Iterable[str | Path]) -> _RowsFromFiles:
    files_seen = 0
    rows_seen = 0
    rows_skipped = 0
    active_rows: list[NCCIPTPRow] = []

    for source_path in files:
        path = Path(source_path)
        files_seen += 1
        for source_name, text in _iter_source_text(path):
            for row in _parse_ptp_text(text, source_name):
                rows_seen += 1
                if row.modifier_indicator == "9":
                    rows_skipped += 1
                    continue
                active_rows.append(row)

    return _RowsFromFiles(
        active=tuple(active_rows),
        summary=NCCIImportSummary(
            files_seen=files_seen,
            rows_seen=rows_seen,
            rows_imported=len(active_rows),
            rows_skipped=rows_skipped,
        ),
    )


def _iter_source_text(path: Path) -> Iterable[tuple[str, str]]:
    if not path.exists():
        raise NCCIImportError(f"NCCI source file does not exist: {path}")

    if path.suffix.lower() == ".zip":
        with zipfile.ZipFile(path) as archive:
            for member in archive.namelist():
                if member.endswith("/"):
                    continue
                suffix = Path(member).suffix.lower()
                if suffix not in {".csv", ".txt", ".xlsx"}:
                    continue
                if suffix == ".xlsx":
                    with archive.open(member) as handle:
                        yield from _iter_xlsx_rows(handle.read(), member)
                else:
                    with archive.open(member) as handle:
                        yield member, _decode_bytes(handle.read())
        return

    if path.suffix.lower() == ".xlsx":
        yield from _iter_xlsx_rows(path.read_bytes(), path.name)
        return

    yield path.name, path.read_text(encoding="utf-8-sig")


def _iter_xlsx_rows(payload: bytes, source_name: str) -> Iterator[tuple[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise NCCIImportError(
            'Excel extraction requires: python -m pip install -e ".[excel]"'
        ) from exc

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        all_rows = list(sheet.iter_rows(values_only=True))
        header_index = _find_excel_header_index(all_rows)
        if header_index is None:
            continue
        header = all_rows[header_index]
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(_normalize_excel_header_row(header))
        for row in all_rows[header_index + 1 :]:
            writer.writerow(["" if cell is None else cell for cell in row])
        yield f"{source_name}:{sheet.title}", output.getvalue()


def _find_excel_header_index(rows: list[tuple[object, ...]]) -> int | None:
    for index, row in enumerate(rows):
        normalized = [_normalize_header(str(cell or "")) for cell in row]
        if "column1" in normalized and "column2" in normalized:
            return index
    return None


def _normalize_excel_header_row(row: tuple[object, ...]) -> list[str]:
    normalized_row: list[str] = []
    for cell in row:
        normalized = _normalize_header(str(cell or ""))
        if normalized == "column1":
            normalized_row.append("Column 1 Code")
        elif normalized == "column2":
            normalized_row.append("Column 2 Code")
        elif normalized == "effective":
            normalized_row.append("Effective Date")
        elif normalized == "deletion":
            normalized_row.append("Deletion Date")
        elif normalized == "modifier":
            normalized_row.append("Modifier Indicator")
        else:
            normalized_row.append("" if cell is None else str(cell))
    return normalized_row


def _decode_bytes(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise NCCIImportError("unable to decode NCCI source file")


def _parse_ptp_text(text: str, source_file: str) -> Iterable[NCCIPTPRow]:
    sample = text[:4096]
    dialect = csv.Sniffer().sniff(sample, delimiters=",|\t")
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if reader.fieldnames is None:
        raise NCCIImportError(f"NCCI file has no header row: {source_file}")

    normalized_headers = {
        _normalize_header(header): header for header in reader.fieldnames
    }
    for row in reader:
        code_a = _value(row, normalized_headers, "column1code", "column1", "code1")
        code_b = _value(row, normalized_headers, "column2code", "column2", "code2")
        modifier_indicator = _value(
            row,
            normalized_headers,
            "modifierindicator",
            "modindicator",
            "ccmi",
        )
        if not code_a or not code_b or modifier_indicator not in {"0", "1", "9"}:
            continue
        yield NCCIPTPRow(
            code_a=code_a,
            code_b=code_b,
            modifier_indicator=modifier_indicator,
            effective_date=_optional_value(
                row, normalized_headers, "effectivedate", "effective"
            ),
            deletion_date=_optional_value(
                row, normalized_headers, "deletiondate", "deletion"
            ),
            rationale=_optional_value(
                row, normalized_headers, "ptpeditrationale", "rationale"
            ),
            source_file=source_file,
        )


def _value(
    row: dict[str, str],
    headers: dict[str, str],
    *candidates: str,
) -> str:
    value = _optional_value(row, headers, *candidates)
    return "" if value is None else value


def _optional_value(
    row: dict[str, str],
    headers: dict[str, str],
    *candidates: str,
) -> str | None:
    for candidate in candidates:
        header = headers.get(candidate)
        if header is not None:
            value = row.get(header, "").strip()
            return value or None
    return None


def _normalize_header(header: str) -> str:
    return "".join(character.lower() for character in header if character.isalnum())


def _blank_to_none(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = value.strip()
    return stripped or None


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Import CMS NCCI PTP edit files.")
    destination = parser.add_mutually_exclusive_group()
    destination.add_argument("--db", help="SQLite reference DB path.")
    destination.add_argument(
        "--postgres-url",
        default=os.getenv("SUPABASE_DB_URL"),
        help="Supabase/Postgres connection URL. Defaults to SUPABASE_DB_URL.",
    )
    destination.add_argument(
        "--extract-csv",
        help="Write normalized CSV instead of loading a database.",
    )
    parser.add_argument(
        "--edit-type",
        choices=["practitioner", "hospital"],
        help="NCCI edit file family being imported.",
    )
    parser.add_argument(
        "--import-version",
        required=False,
        help="CMS version label, for example 2026Q3-v322r0.",
    )
    parser.add_argument(
        "--normalized-csv",
        action="store_true",
        help="Treat input files as normalized CSVs already produced by --extract-csv.",
    )
    parser.add_argument("files", nargs="+", help="CMS ZIP/TXT/CSV files to import.")
    args = parser.parse_args(argv)

    if not args.db and not args.postgres_url and not args.extract_csv:
        parser.error("one of --db, --postgres-url, --extract-csv, or SUPABASE_DB_URL is required")
    if not args.normalized_csv and not args.edit_type:
        parser.error("--edit-type is required unless --normalized-csv is used")
    if not args.normalized_csv and not args.import_version:
        parser.error("--import-version is required unless --normalized-csv is used")

    if args.normalized_csv and args.db:
        summary = load_normalized_ncci_csv_to_sqlite(
            db_path=args.db,
            csv_files=args.files,
        )
    elif args.normalized_csv:
        parser.error("--normalized-csv currently supports SQLite --db loading")
    elif args.extract_csv:
        summary = extract_ncci_ptp_files_to_csv(
            files=args.files,
            output_csv=args.extract_csv,
            edit_type=args.edit_type,
            import_version=args.import_version,
        )
    elif args.postgres_url:
        summary = import_ncci_ptp_files_to_postgres(
            postgres_url=args.postgres_url,
            files=args.files,
            edit_type=args.edit_type,
            import_version=args.import_version,
        )
    else:
        summary = import_ncci_ptp_files(
            db_path=args.db,
            files=args.files,
            edit_type=args.edit_type,
            import_version=args.import_version,
        )
    print(json.dumps(summary.to_dict(), indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
