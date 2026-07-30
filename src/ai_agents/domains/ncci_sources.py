"""CMS NCCI source-file readers and normalizers."""

from __future__ import annotations

import csv
import io
import zipfile
from collections.abc import Iterator
from pathlib import Path
from typing import Iterable

from .ncci_models import NCCIImportError, NCCIImportSummary, NCCIPTPRow, RowsFromFiles


def active_rows_from_files(files: Iterable[str | Path]) -> RowsFromFiles:
    """Read CMS NCCI source files and return active, non-deleted PTP rows."""

    files_seen = 0
    rows_seen = 0
    rows_skipped = 0
    active_rows: list[NCCIPTPRow] = []

    for source_path in files:
        path = Path(source_path)
        files_seen += 1
        for source_name, text in iter_source_text(path):
            for row in parse_ptp_text(text, source_name):
                rows_seen += 1
                if row.modifier_indicator == "9":
                    rows_skipped += 1
                    continue
                active_rows.append(row)

    return RowsFromFiles(
        active=tuple(active_rows),
        summary=NCCIImportSummary(
            files_seen=files_seen,
            rows_seen=rows_seen,
            rows_imported=len(active_rows),
            rows_skipped=rows_skipped,
        ),
    )


def iter_source_text(path: Path) -> Iterable[tuple[str, str]]:
    """Yield text payloads from ZIP, XLSX, CSV, or TXT NCCI sources."""

    if not path.exists():
        raise NCCIImportError(f"NCCI source file does not exist: {path}")

    if path.suffix.lower() == ".zip":
        with zipfile.ZipFile(path) as archive:
            for member in archive.namelist():
                if member.endswith("/"):
                    continue
                suffix = Path(member).suffix.lower()
                if suffix not in {".csv", ".txt", ".xlsx"}:
                    continue
                if suffix == ".xlsx":
                    with archive.open(member) as handle:
                        yield from iter_xlsx_rows(handle.read(), member)
                else:
                    with archive.open(member) as handle:
                        yield member, decode_bytes(handle.read())
        return

    if path.suffix.lower() == ".xlsx":
        yield from iter_xlsx_rows(path.read_bytes(), path.name)
        return

    yield path.name, path.read_text(encoding="utf-8-sig")


def iter_xlsx_rows(payload: bytes, source_name: str) -> Iterator[tuple[str, str]]:
    """Convert NCCI XLSX worksheets into normalized CSV text payloads."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise NCCIImportError(
            'Excel extraction requires: python -m pip install -e ".[excel]"'
        ) from exc

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        all_rows = list(sheet.iter_rows(values_only=True))
        header_index = find_excel_header_index(all_rows)
        if header_index is None:
            continue
        header = all_rows[header_index]
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(normalize_excel_header_row(header))
        for row in all_rows[header_index + 1 :]:
            writer.writerow(["" if cell is None else cell for cell in row])
        yield f"{source_name}:{sheet.title}", output.getvalue()


def find_excel_header_index(rows: list[tuple[object, ...]]) -> int | None:
    """Find the worksheet row containing Column 1/Column 2 headers."""

    for index, row in enumerate(rows):
        normalized = [normalize_header(str(cell or "")) for cell in row]
        if "column1" in normalized and "column2" in normalized:
            return index
    return None


def normalize_excel_header_row(row: tuple[object, ...]) -> list[str]:
    """Map CMS XLSX headers to the canonical text-file header names."""

    normalized_row: list[str] = []
    for cell in row:
        normalized = normalize_header(str(cell or ""))
        if normalized == "column1":
            normalized_row.append("Column 1 Code")
        elif normalized == "column2":
            normalized_row.append("Column 2 Code")
        elif normalized == "effective":
            normalized_row.append("Effective Date")
        elif normalized == "deletion":
            normalized_row.append("Deletion Date")
        elif normalized == "modifier":
            normalized_row.append("Modifier Indicator")
        else:
            normalized_row.append("" if cell is None else str(cell))
    return normalized_row


def decode_bytes(payload: bytes) -> str:
    """Decode CMS text files using the encodings seen in quarterly releases."""

    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise NCCIImportError("unable to decode NCCI source file")


def parse_ptp_text(text: str, source_file: str) -> Iterable[NCCIPTPRow]:
    """Parse one CMS NCCI CSV/TXT text payload into normalized PTP rows."""

    sample = text[:4096]
    dialect = csv.Sniffer().sniff(sample, delimiters=",|\t")
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if reader.fieldnames is None:
        raise NCCIImportError(f"NCCI file has no header row: {source_file}")

    normalized_headers = {normalize_header(header): header for header in reader.fieldnames}
    for row in reader:
        code_a = value(row, normalized_headers, "column1code", "column1", "code1")
        code_b = value(row, normalized_headers, "column2code", "column2", "code2")
        modifier_indicator = value(
            row,
            normalized_headers,
            "modifierindicator",
            "modindicator",
            "ccmi",
        )
        if not code_a or not code_b or modifier_indicator not in {"0", "1", "9"}:
            continue
        yield NCCIPTPRow(
            code_a=code_a,
            code_b=code_b,
            modifier_indicator=modifier_indicator,
            effective_date=optional_value(
                row, normalized_headers, "effectivedate", "effective"
            ),
            deletion_date=optional_value(
                row, normalized_headers, "deletiondate", "deletion"
            ),
            rationale=optional_value(
                row, normalized_headers, "ptpeditrationale", "rationale"
            ),
            source_file=source_file,
        )


def value(row: dict[str, str], headers: dict[str, str], *candidates: str) -> str:
    """Return a normalized required field value or an empty string."""

    result = optional_value(row, headers, *candidates)
    return "" if result is None else result


def optional_value(
    row: dict[str, str],
    headers: dict[str, str],
    *candidates: str,
) -> str | None:
    """Return the first present normalized field value."""

    for candidate in candidates:
        header = headers.get(candidate)
        if header is not None:
            raw_value = row.get(header, "").strip()
            return raw_value or None
    return None


def normalize_header(header: str) -> str:
    """Normalize CMS column names for tolerant cross-release matching."""

    return "".join(character.lower() for character in header if character.isalnum())
